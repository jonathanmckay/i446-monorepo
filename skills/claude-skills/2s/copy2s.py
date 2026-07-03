#!/usr/bin/env python3
"""copy2s.py — copy a month's Neon aggregates into the monthly scorecard.

Source: Neon分v12.2.xlsx › sheet "1分+1s", monthly aggregates at AQ:BA on each
        month's first-week row (col A == <month>.1).
Dest:   scorecard.xlsx › sheet "18-25 2s", row where col A==year AND col B==month.

Positional map (col I "5^1₦" is deprecated → left blank):
    AQ→C  AR→D  AS→E  AT→F  AU→G  AV→H   [I skip]   AW→J  AX→K  AY→L  AZ→M  BA→N

Row numbers are looked up structurally from the on-disk files (stable for past
months), but the VALUES are read LIVE from Excel on Ix, because the saved Neon
file lags the live workbook (a month can still be recomputing after the file was
last saved). All Excel I/O runs on Ix via ix-osa.sh (never local osascript —
OneDrive-conflict policy). Writing opens scorecard.xlsx on Ix if needed and saves.

Usage:
    copy2s.py [MONTH] [--write]
      MONTH   2026-05 | 5 | may | (omitted → previous calendar month)
      --write actually write; without it, dry-run (read live + preview only)
"""
from __future__ import annotations

import datetime as dt
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string as ci

NEON = Path.home() / "OneDrive/vault-excel/Neon分v12.2.xlsx"
SCORECARD = Path.home() / "OneDrive/vault-excel/scorecard.xlsx"
SCORECARD_POSIX = "/Users/mckay/OneDrive/vault-excel/scorecard.xlsx"
SRC_SHEET = "1分+1s"
DST_SHEET = "18-25 2s"
IX_OSA = Path.home() / ".claude/skills/_lib/ix-osa.sh"

SRC_COLS = ["AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA"]
DST_COLS = ["C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N"]  # I skipped
HEADERS = ["sd neon", "neon", "all color", "hcmc/d", "hcb/d", "s897/d",
           "5^0 g/d", "i9", "x2", "-2g", "xk87/d"]
assert len(SRC_COLS) == len(DST_COLS) == len(HEADERS) == 11

MONTHS = {m.lower(): i for i, m in enumerate(
    ["", "jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"]) if i}


def resolve_month(arg: str | None, today: dt.date) -> tuple[int, int]:
    """Return (year, month). Default: previous calendar month."""
    if not arg:
        first = today.replace(day=1)
        prev = first - dt.timedelta(days=1)
        return prev.year, prev.month
    a = arg.strip().lower()
    if "-" in a:  # 2026-05
        y, m = a.split("-")[:2]
        return int(y), int(m)
    if a[:3] in MONTHS:  # may / may2026
        m = MONTHS[a[:3]]
        return today.year, m
    m = int(a)  # bare month number → current tracking year
    return today.year, m


def find_src_row(month: int) -> int:
    wb = openpyxl.load_workbook(NEON, data_only=True, read_only=True)
    s = wb[SRC_SHEET]
    aq = ci("AQ")
    for r in range(2, s.max_row + 1):
        a = s.cell(row=r, column=1).value
        v = s.cell(row=r, column=aq).value
        if isinstance(a, (int, float)) and v not in (None, "") and int(a) == month:
            wb.close()
            return r
    wb.close()
    raise SystemExit(f"ERROR: no month.1 aggregate row for month {month} in {SRC_SHEET}")


def find_dst_row(year: int, month: int) -> tuple[int, bool]:
    """Return (row, already_filled)."""
    wb = openpyxl.load_workbook(SCORECARD, data_only=True, read_only=True)
    s = wb[DST_SHEET]
    for r in range(2, s.max_row + 1):
        y = s.cell(row=r, column=1).value
        m = s.cell(row=r, column=2).value
        if y == year and m == month:
            filled = s.cell(row=r, column=ci("C")).value not in (None, "")
            wb.close()
            return r, bool(filled)
    wb.close()
    raise SystemExit(f"ERROR: no row for {year}-{month:02d} in {DST_SHEET}")


def ix_osa(script: str) -> str:
    r = subprocess.run(["bash", str(IX_OSA)], input=script,
                       capture_output=True, text=True, timeout=90)
    if r.returncode != 0:
        raise SystemExit(f"ix-osa failed (exit {r.returncode}): {r.stderr.strip() or r.stdout.strip()}")
    return r.stdout.strip()


def read_live(src_row: int) -> list[str]:
    reads = "\n".join(
        f'set v{i} to ((value of cell "{c}{src_row}" of srcSheet) as string)'
        for i, c in enumerate(SRC_COLS))
    joined = ' & "|" & '.join(f"v{i}" for i in range(len(SRC_COLS)))
    script = f'''tell application "Microsoft Excel"
    set srcSheet to worksheet "{SRC_SHEET}" of workbook "Neon分v12.2.xlsx"
    {reads}
    return {joined}
end tell'''
    return ix_osa(script).split("|")


def write_live(src_row: int, dst_row: int) -> str:
    reads = "\n".join(
        f'set v{i} to (value of cell "{c}{src_row}" of srcSheet)'
        for i, c in enumerate(SRC_COLS))
    writes = "\n".join(
        f'set value of cell "{c}{dst_row}" of dstSheet to v{i}'
        for i, c in enumerate(DST_COLS))
    script = f'''tell application "Microsoft Excel"
    set srcSheet to worksheet "{SRC_SHEET}" of workbook "Neon分v12.2.xlsx"
    {reads}
    set scOpen to false
    repeat with w in workbooks
        if (name of w) is "scorecard.xlsx" then set scOpen to true
    end repeat
    if not scOpen then
        open workbook workbook file name ((POSIX file "{SCORECARD_POSIX}") as string)
    end if
    set dstWB to workbook "scorecard.xlsx"
    set dstSheet to worksheet "{DST_SHEET}" of dstWB
    {writes}
    save dstWB
    return "OK"
end tell'''
    return ix_osa(script)


def main():
    args = [a for a in sys.argv[1:]]
    write = "--write" in args
    args = [a for a in args if a != "--write"]
    month_arg = args[0] if args else None

    today = dt.date.today()
    year, month = resolve_month(month_arg, today)
    src_row = find_src_row(month)
    dst_row, filled = find_dst_row(year, month)
    values = read_live(src_row)

    mname = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"][month]
    print(f"{mname} {year}  |  src 1分+1s row {src_row} (AQ:BA)  →  dst 18-25 2s row {dst_row} (C:N, I blank)")
    print(f"{'metric':<10} {'src':>5} {'live value'}")
    for h, c, v in zip(HEADERS, DST_COLS, values):
        print(f"  {h:<10} {c:>3}  {v}")
    if filled:
        print(f"\n⚠ dst row {dst_row} column C is already non-blank — writing will overwrite.")

    if not write:
        print("\n(dry run — pass --write to apply)")
        return

    write_live(src_row, dst_row)
    print(f"\n✓ wrote {mname} {year} → scorecard 18-25 2s row {dst_row} (saved on Ix).")


if __name__ == "__main__":
    main()
