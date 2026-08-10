"""Regression tests for 0t-fast.py."""
from __future__ import annotations

import importlib.util
import sys
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch

# Load the hyphenated module
_PATH = Path(__file__).parent / "0t-fast.py"
_SPEC = importlib.util.spec_from_file_location("zerot_fast", _PATH)
zerot_fast = importlib.util.module_from_spec(_SPEC)
sys.modules["zerot_fast"] = zerot_fast
_SPEC.loader.exec_module(zerot_fast)


def test_tag_columns_match_live_headers():
    """0n headers are AU='1+', AV='-1', AW='-2', AX='-3', AS='其他人'. The -1 tag
    must write to AV and -2 to AW. Regression: a '1+' column was inserted at AU,
    so an old map of -1→AU duplicated -1 points into the 1+ column (AU + AV)."""
    assert zerot_fast.TAG_COLUMNS["-1"] == "AV"
    assert zerot_fast.TAG_COLUMNS["-2"] == "AW"
    assert zerot_fast.TAG_COLUMNS["-3"] == "AX"
    assert zerot_fast.TAG_COLUMNS["其他人"] == "AS"
    assert "xk87" not in zerot_fast.TAG_COLUMNS, (
        "AZ ('∑xk87') is a live SUM formula, not a tag-total target; "
        "writing to it clobbers the formula"
    )


def test_tag_columns_agree_with_daemon():
    """0t-fast and build-order-daemon both write 0n tag columns; they must not
    drift (the original bug was 0t-fast lagging the daemon after a column shift)."""
    daemon_path = Path(__file__).resolve().parents[2] / "scripts" / "build-order-daemon.py"
    spec = importlib.util.spec_from_file_location("bod_daemon", daemon_path)
    daemon = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(daemon)
    shared = set(zerot_fast.TAG_COLUMNS) & set(daemon.TOGGL_TAG_COLS)
    assert shared, "expected overlapping tag keys between the two maps"
    for tag in shared:
        assert zerot_fast.TAG_COLUMNS[tag] == daemon.TOGGL_TAG_COLS[tag], (
            f"{tag}: 0t-fast={zerot_fast.TAG_COLUMNS[tag]} "
            f"daemon={daemon.TOGGL_TAG_COLS[tag]}")


def test_compute_tag_minutes_excludes_sleep_from_minus3():
    """睡觉 (sleep) carries the -3 tag but is tracked in column D. It must NOT be
    summed into the -3/AX tag column, or AX reads as a full night of sleep
    (regression 2026-06-28: AX=439)."""
    sleep = {"duration": 400 * 60, "tags": ["-3"], "project_id": zerot_fast.SLEEP_PROJECT_ID}
    media = {"duration": 30 * 60, "tags": ["-3"], "project_id": 109932707}  # 新闻/hcmc
    with patch.object(zerot_fast, "get_toggl_entries", return_value=[sleep, media]):
        tag_totals, _ = zerot_fast.compute_tag_minutes(date(2026, 6, 27), date(2026, 6, 28))
    assert tag_totals.get("-3") == 30, f"AX must exclude sleep; got {tag_totals.get('-3')}"


def test_daemon_compute_toggl_totals_excludes_sleep_from_AX():
    """The daemon writes today's AX; it must also exclude sleep from -3."""
    daemon_path = Path(__file__).resolve().parents[2] / "scripts" / "build-order-daemon.py"
    spec = importlib.util.spec_from_file_location("bod_daemon2", daemon_path)
    daemon = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(daemon)
    sleep = {"duration": 400 * 60, "tags": ["-3"], "project_id": daemon.SLEEP_PROJECT_ID}
    media = {"duration": 25 * 60, "tags": ["-3"], "project_id": 109932707}
    with patch.object(daemon, "_toggl_get", return_value=[sleep, media]):
        totals = daemon.compute_toggl_totals(date(2026, 6, 28))
    assert totals.get("AX") == 25, f"daemon AX must exclude sleep; got {totals.get('AX')}"


def test_mark_night_hcmc_targets_entry_day_row():
    """night hcmc minutes must be logged to the date the entry occurred
    (yesterday), not today. Rewritten 2026-07-24: the old version routed
    through did-fast, whose 0n path refuses past dates — so the call site
    passed `today`, and a backfill run (`0t-fast.py 2026-07-22`) stamped
    7/22's detected minutes onto 7/24's row, clobbering a manual value.
    Now it's a direct row-targeted write with an empty-cell guard."""
    captured = {}

    class _FakeRes:
        returncode = 0
        stdout = "OK: night hcmc=35 row=180"
        stderr = ""

    def _fake_ix_run(script, timeout=30.0):
        captured["script"] = script
        return _FakeRes()

    with patch.object(zerot_fast, "ix_run", side_effect=_fake_ix_run):
        out = zerot_fast.mark_night_hcmc(35, date(2026, 5, 6))

    script = captured["script"]
    assert "if m = 5 and d = 6 then" in script, (
        "write must locate the ENTRY day's row, not today's")
    # Manual /did values always win: only an empty/zero cell may be written.
    assert 'if prev = "" or prev = "0" then' in script
    assert "SKIPPED: manual value" in script
    assert out == {"write": "OK: night hcmc=35 row=180"}


def test_main_passes_yesterday_to_night_hcmc():
    """The call site must hand mark_night_hcmc the entry's day (yesterday),
    never `today` (the 2026-07-24 backfill clobber)."""
    src = _PATH.read_text()
    assert "mark_night_hcmc(night_hcmc, yesterday)" in src
    assert "mark_night_hcmc(night_hcmc, today)" not in src


def test_tag_and_project_minutes_only_count_target_day():
    """All tag and project minute totals must only sum the target day (yesterday),
    not both days. Summing today's entries inflates yesterday's row."""
    yesterday = date(2026, 5, 16)
    today = date(2026, 5, 17)

    yesterday_entries = [
        {"duration": 18000, "tags": ["-3", "xk87"], "project_id": 163129781},  # 300min
    ]
    today_entries = [
        {"duration": 24000, "tags": ["-3", "xk87"], "project_id": 163129781},  # 400min
    ]

    def fake_entries(d):
        if d == yesterday:
            return yesterday_entries
        return today_entries

    with patch.object(zerot_fast, "get_toggl_entries", side_effect=fake_entries):
        tag_totals, proj_totals = zerot_fast.compute_tag_minutes(yesterday, today)

    # Tags must only count yesterday
    assert tag_totals["-3"] == 300, f"-3 tag should be 300 (yesterday only), got {tag_totals['-3']}"
    # xk87 isn't a tracked tag column (AZ is a live SUM formula, not a write target)
    assert "xk87" not in tag_totals, f"xk87 should not be tracked, got {tag_totals}"
    # Projects should be empty (no project_id in PROJECT_COLUMNS)
    assert not proj_totals, f"proj_totals should be empty, got {proj_totals}"


def test_write_tag_minutes_uses_absolute_overwrite():
    """Tag/project minute totals in 0n must use absolute overwrite (set value of),
    not append. These are recalculated totals from Toggl, not incremental points.
    Appending caused ballooning values when 0t or the daemon ran multiple times."""
    import ast
    source = Path(__file__).parent.joinpath("0t-fast.py").read_text()
    tree = ast.parse(source)

    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef) and node.name == "write_tag_minutes":
            body_src = ast.get_source_segment(source, node)
            assert "set value of" in body_src, (
                "write_tag_minutes must use absolute overwrite (set value of) for Toggl minute totals"
            )
            assert "oldVal" not in body_src, (
                "write_tag_minutes must NOT append to old formula — these are absolute totals"
            )
            return
    raise AssertionError("write_tag_minutes function not found")


def test_refresh_points_cache_includes_block_data(tmp_path):
    """Regression (2026-07-19): refresh_points_cache() only read the P:Y domain
    -total columns (COLS), never the G:O per-block (地支 卯..亥) columns, and
    never wrote a "__block__" key. dashboard.py's load_points_all() reads this
    SAME .points-cache.json file, and _build_block_chart_data() needs the
    "__block__" sub-dict to render Points/Block at all — so every /0t run (it
    calls refresh_points_cache() every morning) silently wiped block data,
    making Points/Block show empty even though the daemon writes real values
    to G:O in Neon. Fix: also read G:O into a "__block__" sub-dict, matching
    the shape dashboard.py's own xlwings fallback path already produces."""
    import json as _json
    import openpyxl as _openpyxl
    import time as _time

    yesterday = date.today() - timedelta(days=1)

    # Row layout (0-indexed): [0]=A, [1]=B(date), [6]=G(卯) .. [14]=O(亥),
    # [15]=P(-1₦) .. [24]=Y(社) — matches COLS/BLOCK_COLS' 1-indexed offsets.
    row = [None] * 25
    row[1] = yesterday
    row[6] = 6     # G — 卯
    row[9] = 13    # J — 午
    row[15] = 10   # P — -1₦ (a domain total, sanity check COLS still works)

    class _FakeSheet:
        def iter_rows(self, min_row, values_only):
            yield tuple(row)

    class _FakeWorkbook:
        def __getitem__(self, name):
            assert name == "0分"
            return _FakeSheet()

        def close(self):
            pass

    fake_cache = tmp_path / ".points-cache.json"
    with patch.object(_openpyxl, "load_workbook", return_value=_FakeWorkbook()), \
         patch.object(zerot_fast, "ix_run", return_value=None), \
         patch.object(zerot_fast, "POINTS_CACHE", fake_cache), \
         patch.object(_time, "sleep", return_value=None):
        zerot_fast.refresh_points_cache()

    cache = _json.loads(fake_cache.read_text())
    day = cache[yesterday.isoformat()]
    assert day.get("__block__") == {"卯": 6, "午": 13}, (
        f"expected __block__ with 卯/午 values, got {day.get('__block__')}"
    )
    assert day.get("-1₦") == 10, "domain-total columns (COLS) must still work"


def test_marks_done_then_refreshes_dtd_cache():
    """/0t records 0t in completed-today via mark_done(), but dtd only reloads on a
    cache mtime change — so 0t-fast must run did-fast --refresh-cache after marking
    done, else 0t lingers on the dtd list (regression 2026-06-30)."""
    import ast
    src = (Path(__file__).parent / "0t-fast.py").read_text()
    main = next(n for n in ast.walk(ast.parse(src))
                if isinstance(n, ast.FunctionDef) and n.name == "main")
    body = ast.get_source_segment(src, main)
    assert '"--refresh-cache"' in body, "0t-fast main() must refresh the dtd cache"
    # and it must come AFTER mark_done() so completed-today is already written
    assert body.index("mark_done()") < body.index('"--refresh-cache"'), \
        "refresh must follow mark_done()"


# ── Sleep dock ────────────────────────────────────────────────────────────────

def _te(start_local: str, dur_min: int, project_id: int | None = None,
        desc: str = "") -> dict:
    """Synthetic Toggl entry with a PDT-offset ISO start."""
    return {"id": hash((start_local, desc)), "start": start_local + "-07:00",
            "duration": dur_min * 60, "project_id": project_id,
            "description": desc}


_Y = date(2026, 8, 9)
_S = date(2026, 8, 10)
_SLEEP = zerot_fast.SLEEP_PROJECT_ID
_HCMC = zerot_fast.HCMC_PROJECT_ID


def _dock_with(entries):
    with patch.object(zerot_fast, "gather_entries_local", return_value=entries):
        return zerot_fast.compute_sleep_dock(_Y, _S)


def test_dock_abandoned_attempt_does_not_count():
    """Real night 2026-08-08→09: fall asleep 21:00 abandoned (youtube after),
    second attempt 22:00 chains into 睡觉 22:30. Bedtime = 22:00, dock 0."""
    r = _dock_with([
        _te("2026-08-09T21:00:00", 15, _HCMC, "fall asleep"),
        _te("2026-08-09T21:15:00", 45, None, "youtube"),
        _te("2026-08-09T22:00:00", 30, _HCMC, "fall asleep"),
        _te("2026-08-09T22:30:00", 89, _SLEEP, "睡觉"),
    ])
    assert r["basis"] == "fall asleep"
    assert r["late_minutes"] == 0
    assert r["dock"] == 0


def test_dock_abandoned_only_attempt_falls_back_to_sleep_start():
    """fall asleep 21:30, up again, 睡觉 23:45 with no second attempt:
    the 21:30 attempt must NOT suppress the dock."""
    r = _dock_with([
        _te("2026-08-09T21:30:00", 15, _HCMC, "fall asleep"),
        _te("2026-08-09T23:45:00", 14, _SLEEP, "睡觉"),
    ])
    assert r["basis"] == "睡觉"
    assert r["late_minutes"] == 105
    assert r["dock"] == 26.25


def test_dock_chains_through_multiple_fall_asleeps():
    r = _dock_with([
        _te("2026-08-09T22:10:00", 20, _HCMC, "fall asleep"),
        _te("2026-08-09T22:30:00", 30, _HCMC, "fall asleep"),
        _te("2026-08-09T23:00:00", 59, _SLEEP, "睡觉"),
    ])
    assert r["bedtime"].endswith("22:10")
    assert r["late_minutes"] == 10


def test_dock_cross_midnight_is_capped():
    r = _dock_with([_te("2026-08-10T00:30:00", 300, _SLEEP, "睡觉")])
    assert r["late_minutes"] == 150
    assert r["dock"] == 30.0   # 150 * 0.25 = 37.5, capped


def test_dock_midnight_wakeup_fall_asleep_ignored():
    """A 3am re-fall-asleep after a night waking must not become the bedtime."""
    r = _dock_with([
        _te("2026-08-09T21:30:00", 30, _HCMC, "fall asleep"),
        _te("2026-08-09T22:00:00", 119, _SLEEP, "睡觉"),
        _te("2026-08-10T03:00:00", 20, _HCMC, "fall asleep"),
        _te("2026-08-10T03:20:00", 200, _SLEEP, "睡觉"),
    ])
    assert r["bedtime"].endswith("21:30")
    assert r["dock"] == 0


def test_dock_in_bed_before_22_is_zero():
    r = _dock_with([_te("2026-08-09T21:15:00", 400, _SLEEP, "睡觉")])
    assert r["late_minutes"] == 0
    assert r["dock"] == 0


def test_dock_no_entries_returns_none():
    assert _dock_with([_te("2026-08-09T21:00:00", 30, None, "reading")]) is None


def test_dock_daytime_nap_not_a_bedtime():
    """A 12:50pm weekend nap chain is outside the <14:00 guard only by luck of
    hour; the first-睡觉-of-night rule keeps the real bedtime authoritative."""
    r = _dock_with([
        _te("2026-08-09T22:30:00", 60, _SLEEP, "睡觉"),
        _te("2026-08-10T12:50:00", 10, _HCMC, "fall asleep"),
        _te("2026-08-10T13:00:00", 40, _SLEEP, "睡觉"),
    ])
    assert r["bedtime"].endswith("22:30")


def test_dock_write_goes_through_neon_excel_not_osascript():
    """0分 writes are banned from raw AppleScript; the dock must use
    lib/neon/excel (daemon + ledger)."""
    src = _PATH.read_text()
    dock_fn = src.split("def write_sleep_dock", 1)[1].split("\ndef ", 1)[0]
    assert "neon_excel.append" in dock_fn
    assert "ix_run" not in dock_fn


def test_dock_write_fails_closed_when_ledger_unreachable():
    with patch.object(zerot_fast, "_dock_already_logged", return_value=None), \
         patch.object(zerot_fast.neon_excel, "append") as mock_append:
        r = zerot_fast.write_sleep_dock(7.5, _Y)
    assert "error" in r
    mock_append.assert_not_called()


def test_dock_write_skips_when_already_in_ledger():
    with patch.object(zerot_fast, "_dock_already_logged", return_value=True), \
         patch.object(zerot_fast.neon_excel, "append") as mock_append:
        r = zerot_fast.write_sleep_dock(7.5, _Y)
    assert "skipped" in r["write"]
    mock_append.assert_not_called()


def test_dock_write_appends_negative_term():
    with patch.object(zerot_fast, "_dock_already_logged", return_value=False), \
         patch.object(zerot_fast.neon_excel, "append",
                      return_value={"ok": True, "after_value": "139.0"}) as mock_append:
        r = zerot_fast.write_sleep_dock(7.5, _Y)
    assert "-7.5" in r["write"]
    kwargs = mock_append.call_args.kwargs
    assert mock_append.call_args.args == ("0分", "W")
    assert kwargs["value"] == "-7.5"
    assert kwargs["date"] == "8/9"
    assert kwargs["src"] == "0t sleep-dock"


# ── Media audit ───────────────────────────────────────────────────────────────

import importlib.util as _ilu

_MA_PATH = Path(__file__).parent / "media_audit.py"
_MA_SPEC = _ilu.spec_from_file_location("media_audit_test", _MA_PATH)
ma = _ilu.module_from_spec(_MA_SPEC)
_MA_SPEC.loader.exec_module(ma)

from datetime import datetime, timedelta, timezone as _tz


def _aw_ev(start_utc: str, dur_s: int, **data) -> dict:
    return {"timestamp": start_utc, "duration": dur_s, "data": data}


def test_classify_media_titles():
    assert ma._classify("Google Chrome — MKBHD - YouTube") == "YouTube"
    assert ma._classify("Audible") == "Audible"
    assert ma._classify("Obsidian — notes.md") is None


def test_overlap_clips_to_active_intervals():
    t0 = datetime(2026, 8, 9, 20, 0, tzinfo=_tz.utc)
    active = [(t0, t0 + timedelta(minutes=10))]
    ev = (t0 + timedelta(minutes=5), t0 + timedelta(minutes=30))
    assert ma._overlap_seconds(ev, active) == 300  # only 5 min overlap


def test_media_audit_flags_large_gap():
    with patch.object(ma, "media_minutes_from_aw", return_value={"YouTube": 85.0}), \
         patch.object(ma, "media_minutes_from_screentime", return_value=None):
        r = ma.media_audit(date(2026, 8, 9), [], lambda e: None)
    assert r["passive_total_min"] == 85
    assert r["toggl_media_min"] == 0
    assert r["flagged"] is True
    assert any("Screen Time" in n for n in r["notes"])


def test_media_audit_no_flag_when_toggl_covers_it():
    entries = [{"project_id": 109932707, "duration": 80 * 60, "start": "x"}]
    ldt = lambda e: datetime(2026, 8, 9, 12, 0, tzinfo=ma.LOCAL_TZ)
    with patch.object(ma, "media_minutes_from_aw", return_value={"YouTube": 85.0}), \
         patch.object(ma, "media_minutes_from_screentime", return_value=None):
        r = ma.media_audit(date(2026, 8, 9), entries, ldt)
    assert r["toggl_media_min"] == 80
    assert r["flagged"] is False


def test_media_audit_merges_sources_by_max_not_sum():
    with patch.object(ma, "media_minutes_from_aw", return_value={"YouTube": 60.0}), \
         patch.object(ma, "media_minutes_from_screentime", return_value={"YouTube": 45.0, "Audible": 30.0}):
        r = ma.media_audit(date(2026, 8, 9), [], lambda e: None)
    assert r["passive"]["YouTube"] == 60.0   # max, not 105
    assert r["passive"]["Audible"] == 30.0
    assert r["passive_total_min"] == 90
