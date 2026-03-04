import datetime
import json
import os

from openpyxl import load_workbook

from .config import NEON_EXCEL_PATH

# Known category codes that appear in column A of the 1g sheet.
WEEKLY_GOAL_CATEGORIES = [
    "i8", "m5c7", "hcmp", "hcb", "g245", "hci", "hcmc", "s897",
]

# Column mapping for the 1g sheet (1-indexed).
_1G = {
    "category": 1,   # A
    "goal": 4,        # D
    "minutes": 5,     # E
    "bonus": 6,       # F
    "pct_done": 7,    # G
    "score": 9,       # I
    "daily_start": 11,  # K
    "daily_end": 17,    # Q
}

DAY_LABELS = ["T", "W", "Th", "F", "Sa", "Su", "M"]


class NeonHandler:
    def __init__(self, file_path=None):
        self.file_path = file_path or NEON_EXCEL_PATH

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _open(self, data_only=False):
        """Open the workbook fresh (no long-lived handle — OneDrive safe)."""
        return load_workbook(self.file_path, data_only=data_only)

    def _save(self, wb):
        wb.save(self.file_path)

    # ------------------------------------------------------------------
    # 1g — Weekly Goals
    # ------------------------------------------------------------------

    def get_weekly_goals(self):
        """Parse the 1g sheet into structured weekly-goals data."""
        wb = self._open(data_only=True)
        ws = wb["1g"]

        tldr = ws.cell(row=1, column=1).value or ""
        week_label = ws.cell(row=2, column=1).value or ""

        categories = []
        current_cat = None
        current_goals = []

        for row in range(3, ws.max_row + 1):
            a_val = ws.cell(row=row, column=_1G["category"]).value
            d_val = ws.cell(row=row, column=_1G["goal"]).value

            a_str = str(a_val).strip() if a_val else ""
            if a_str in WEEKLY_GOAL_CATEGORIES:
                if current_cat is not None:
                    categories.append({"code": current_cat, "goals": current_goals})
                current_cat = a_str
                current_goals = []
                continue

            if not d_val or not str(d_val).strip():
                continue
            if current_cat is None:
                continue

            goal = {
                "row": row,
                "text": str(d_val).strip(),
                "minutes": ws.cell(row=row, column=_1G["minutes"]).value,
                "bonus": ws.cell(row=row, column=_1G["bonus"]).value,
                "pct_done": ws.cell(row=row, column=_1G["pct_done"]).value,
                "score": ws.cell(row=row, column=_1G["score"]).value,
                "daily": {},
            }
            for i, day in enumerate(DAY_LABELS):
                col = _1G["daily_start"] + i
                goal["daily"][day] = ws.cell(row=row, column=col).value
            current_goals.append(goal)

        if current_cat is not None:
            categories.append({"code": current_cat, "goals": current_goals})

        # Find summary totals row (first row after all goals with a value in col F).
        totals = {}
        if categories:
            all_goal_rows = [
                g["row"] for cat in categories for g in cat["goals"]
            ]
            last_goal_row = max(all_goal_rows) if all_goal_rows else 42
            for row in range(last_goal_row + 1, last_goal_row + 10):
                f_val = ws.cell(row=row, column=_1G["bonus"]).value
                if f_val is not None:
                    totals = {
                        "bonus": f_val,
                        "pct_done": ws.cell(row=row, column=_1G["pct_done"]).value,
                        "score": ws.cell(row=row, column=_1G["score"]).value,
                    }
                    break

        wb.close()
        return {
            "tldr": tldr,
            "week_label": week_label,
            "categories": categories,
            "totals": totals,
        }

    def update_weekly_goal(self, category, goal_index, field, value):
        """Update a single field of an existing weekly goal.

        category: domain code (e.g. "g245")
        goal_index: 0-based index within the category
        field: "text", "minutes", "bonus", "pct_done",
               or "daily_T", "daily_W", etc.
        value: new value to write
        """
        goals_data = self.get_weekly_goals()
        cat_data = next(
            (c for c in goals_data["categories"] if c["code"] == category), None
        )
        if cat_data is None:
            avail = [c["code"] for c in goals_data["categories"]]
            return f"Category '{category}' not found. Available: {avail}"
        if goal_index < 0 or goal_index >= len(cat_data["goals"]):
            return (
                f"Goal index {goal_index} out of range "
                f"('{category}' has {len(cat_data['goals'])} goals)"
            )

        target_row = cat_data["goals"][goal_index]["row"]

        # Resolve column.
        if field.startswith("daily_"):
            day = field.split("_", 1)[1]
            if day not in DAY_LABELS:
                return f"Unknown day '{day}'. Use one of: {DAY_LABELS}"
            col = _1G["daily_start"] + DAY_LABELS.index(day)
        elif field in _1G:
            col = _1G[field]
        else:
            return (
                f"Unknown field '{field}'. "
                "Use: text, minutes, bonus, pct_done, or daily_T/W/Th/F/Sa/Su/M"
            )

        # Coerce types.
        if field in ("minutes", "bonus", "score") and value is not None:
            try:
                value = int(value)
            except (ValueError, TypeError):
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
        if field == "pct_done" and isinstance(value, str) and value.endswith("%"):
            try:
                value = float(value.rstrip("%")) / 100
            except ValueError:
                pass
        if field.startswith("daily_") and value is not None:
            try:
                value = int(value)
            except (ValueError, TypeError):
                pass

        wb = self._open(data_only=False)
        ws = wb["1g"]
        ws.cell(row=target_row, column=col).value = value
        self._save(wb)
        wb.close()
        return f"Updated {category}[{goal_index}].{field} = {value} (row {target_row})"

    def add_weekly_goal(self, category, goal_text, minutes=0, bonus=0):
        """Add a new goal row at the end of a category block in the 1g sheet."""
        goals_data = self.get_weekly_goals()
        cat_data = next(
            (c for c in goals_data["categories"] if c["code"] == category), None
        )
        if cat_data is None:
            return f"Category '{category}' not found."

        if cat_data["goals"]:
            insert_row = cat_data["goals"][-1]["row"] + 1
        else:
            # Find the category header row.
            wb_tmp = self._open(data_only=True)
            ws_tmp = wb_tmp["1g"]
            insert_row = None
            for row in range(3, ws_tmp.max_row + 1):
                a_val = ws_tmp.cell(row=row, column=1).value
                if a_val and str(a_val).strip() == category:
                    insert_row = row + 1
                    break
            wb_tmp.close()
            if insert_row is None:
                return f"Could not find category header for '{category}'."

        wb = self._open(data_only=False)
        ws = wb["1g"]
        ws.insert_rows(insert_row)
        ws.cell(row=insert_row, column=_1G["goal"]).value = goal_text
        if minutes:
            ws.cell(row=insert_row, column=_1G["minutes"]).value = minutes
        if bonus:
            ws.cell(row=insert_row, column=_1G["bonus"]).value = bonus
        ws.cell(row=insert_row, column=_1G["pct_done"]).value = 0
        self._save(wb)
        wb.close()
        return f"Added goal '{goal_text}' to {category} at row {insert_row}"

    def set_weekly_tldr(self, text):
        """Update the tl;dr summary in row 1 of the 1g sheet."""
        wb = self._open(data_only=False)
        ws = wb["1g"]
        ws.cell(row=1, column=1).value = text
        self._save(wb)
        wb.close()
        return f"Updated weekly tl;dr."

    # ------------------------------------------------------------------
    # General / Legacy (0s, 0₦ sheets)
    # ------------------------------------------------------------------

    def find_date_row(self, sheet_name="0s", target_date=None):
        if target_date is None:
            target_date = datetime.datetime.now().date()

        if sheet_name == "0s":
            try:
                map_path = os.path.join(os.path.dirname(__file__), "date_map.json")
                if os.path.exists(map_path):
                    with open(map_path) as f:
                        date_map = json.load(f)
                    target_str = target_date.strftime("%Y-%m-%d")
                    if target_str in date_map:
                        return date_map[target_str]
            except Exception:
                pass

        wb = self._open(data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]
        col_idx = 3 if sheet_name == "0₦" else 2
        for r in range(1, 1000):
            val = ws.cell(row=r, column=col_idx).value
            if isinstance(val, datetime.datetime) and val.date() == target_date:
                wb.close()
                return r
        wb.close()
        return None

    def get_headers(self, sheet_name="0s", row=1):
        wb = self._open(data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[sheet_name]
        headers = {}
        for c in range(1, 60):
            val = ws.cell(row=row, column=c).value
            if val:
                headers[str(val).strip()] = c
        wb.close()
        return headers

    def get_row_values(self, sheet_name, row):
        wb = self._open(data_only=True)
        ws = wb[sheet_name]
        values = {}
        for c in range(1, 60):
            values[c] = ws.cell(row=row, column=c).value
        wb.close()
        return values

    def update_cell(self, sheet_name, row, col_idx, value):
        wb = self._open(data_only=False)
        ws = wb[sheet_name]
        ws.cell(row=row, column=col_idx).value = value
        self._save(wb)
        wb.close()
        return True

    def list_sheets(self):
        wb = self._open(data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
