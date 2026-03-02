import pandas as pd
from openpyxl import load_workbook

def inspect_0fen_row400():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True) # Read values
    sheet = wb['0分']
    
    print("--- 0分 Sheet Rows 390-410 ---")
    data = []
    for r in range(390, 410):
        row_data = [f"R{r}"]
        for c in range(1, 10): 
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)

    df = pd.DataFrame(data, columns=["Row"] + [f"C{i}" for i in range(1, 10)])
    print(df.to_markdown(index=False))

if __name__ == "__main__":
    inspect_0fen_row400()
