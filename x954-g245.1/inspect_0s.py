import pandas as pd
from openpyxl import load_workbook

def inspect_0s():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0s']
    
    print("--- 0s Sheet Inspection ---")
    # specific cells to check hypothesis: Rows = Days
    # Check A2 to A20
    # Check C2 to C20 (DOW?)
    
    data = []
    for r in range(30, 70):
        row_data = []
        for c in range(1, 15): # A to N
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)

    df = pd.DataFrame(data, columns=[f"C{i}" for i in range(1, 15)])
    print(df.to_markdown())

if __name__ == "__main__":
    inspect_0s()
