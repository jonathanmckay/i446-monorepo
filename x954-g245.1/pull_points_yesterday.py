import pandas as pd
from openpyxl import load_workbook
import datetime

def pull_points():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0₦']
    
    # Headers in Row 1
    headers = {}
    for c in range(4, 30): # Start from Col 4 (Col 1-3 are Date/Meta)
        val = sheet.cell(row=1, column=c).value
        if val:
            headers[c] = val
            
    # Target Row for Yesterday (2026-02-05)
    # Based on inspection: Row 5 is Jan 4. 2026-02-05 is +32 days.
    # Row = 5 + 32 = 37.
    # Let's verify date in Col 3
    target_row = 37
    date_val = sheet.cell(row=target_row, column=3).value
    print(f"Checking Row {target_row}. Date found: {date_val}")
    
    print("\n--- Points for Yesterday ---")
    points = []
    for c, name in headers.items():
        val = sheet.cell(row=target_row, column=c).value
        if val and isinstance(val, (int, float)) and val != 0:
            points.append((name, val))
            
    # Formatted Output
    df = pd.DataFrame(points, columns=["Category", "Points"])
    print(df.to_markdown(index=False))

if __name__ == "__main__":
    pull_points()
