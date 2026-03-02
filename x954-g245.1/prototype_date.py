from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

def find_today_cell(file_path):
    print(f"Loading {file_path}...")
    wb = load_workbook(file_path, data_only=True)
    
    # Check '0s' sheet first as it appeared to have dates
    sheet = wb['0s']
    
    # Iterate through first few rows/cols to find dates
    print("Searching for dates...")
    
    # Just printing first 20 cells in first row to see format
    for i in range(1, 20):
        cell_val = sheet.cell(row=1, column=i).value
        print(f"Row 1, Col {i}: {cell_val} (Type: {type(cell_val)})")

    # printing first 20 cells in first col
    for i in range(1, 20):
        cell_val = sheet.cell(row=i, column=1).value
        print(f"Row {i}, Col 1: {cell_val} (Type: {type(cell_val)})")

if __name__ == "__main__":
    find_today_cell('neon_agent_copy.xlsx')
