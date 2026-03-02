import keyring
import base64
import json
from urllib import request
from openpyxl import load_workbook

token = keyring.get_password('neon-prayer-merge', 'TOGGL_API_TOKEN')
creds = base64.b64encode(f'{token}:api_token'.encode()).decode()

# Check Toggl entries for Feb 10
req = request.Request(
    'https://api.track.toggl.com/api/v9/me/time_entries?start_date=2026-02-10&end_date=2026-02-11',
    headers={'Authorization': f'Basic {creds}'}
)
with request.urlopen(req) as resp:
    entries = json.loads(resp.read().decode())
    prayer_entries = [e for e in entries if 'الفاتحة' in e.get('description', '')]
    print(f'Prayer entries in Toggl for Feb 10: {len(prayer_entries)}')
    for e in prayer_entries:
        print(f"  - {e.get('description')} (ID: {e.get('id')})")

# Check Excel
print()
excel_path = '/Users/mckay/Library/CloudStorage/OneDrive-Personal/Neon分v11.xlsx'
wb = load_workbook(excel_path, data_only=False)
sheet = wb['0₦']
cell_formula = sheet.cell(row=42, column=39).value
wb.close()

wb2 = load_workbook(excel_path, data_only=True)
sheet2 = wb2['0₦']
cell_value = sheet2.cell(row=42, column=39).value
wb2.close()

print(f'Excel AM42 (row 42, col 39):')
print(f'  Formula: {cell_formula}')
print(f'  Value: {cell_value}')
