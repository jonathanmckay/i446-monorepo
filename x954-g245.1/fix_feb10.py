#!/usr/bin/env python3
"""One-time fix to properly merge Feb 10 prayers"""
import keyring
import base64
import json
import re
from urllib import request
from openpyxl import load_workbook

token = keyring.get_password('neon-prayer-merge', 'TOGGL_API_TOKEN')
workspace_id = keyring.get_password('neon-prayer-merge', 'TOGGL_WORKSPACE_ID')
project_id = keyring.get_password('neon-prayer-merge', 'TOGGL_PROJECT_ID')
creds = base64.b64encode(f'{token}:api_token'.encode()).decode()

excel_path = '/Users/mckay/Library/CloudStorage/OneDrive-Personal/Neon分v11.xlsx'

# 1. Get Toggl prayer entries for Feb 10
print("1. Fetching Toggl entries for Feb 10...")
req = request.Request(
    'https://api.track.toggl.com/api/v9/me/time_entries?start_date=2026-02-10&end_date=2026-02-11',
    headers={'Authorization': f'Basic {creds}'}
)
with request.urlopen(req) as resp:
    entries = json.loads(resp.read().decode())
    prayer_entries = [e for e in entries if 'الفاتحة' in e.get('description', '')]

# Count single prayers (not 'Nx الفاتحة' format we created)
single_prayers = [e for e in prayer_entries if not re.match(r'\d+x', e.get('description', ''))]
toggl_count = len(single_prayers)
print(f"   Found {toggl_count} single prayer entries to merge")

# Pick the first single prayer entry to update
entry_to_update = single_prayers[0] if single_prayers else None
print(f"   Entry to update: {entry_to_update.get('id') if entry_to_update else None}")

# 2. Read current Excel formula
print("\n2. Updating Excel AM42...")
wb = load_workbook(excel_path, data_only=False)
sheet = wb['0₦']
current_formula = sheet.cell(row=42, column=39).value
print(f"   Current: {current_formula}")

# User said original was 27, plus 8 from toggl = 35
# We have 10 single prayers, so: 27 + 10 = 37
# Actually user said 8 prayers in toggl -> 35 total, so original = 27
new_formula = f'=27+{toggl_count}'
sheet.cell(row=42, column=39).value = new_formula
wb.save(excel_path)
wb.close()
print(f"   Updated to: {new_formula}")

# 3. Calculate total
total = 27 + toggl_count
print(f"\n3. Total prayers: {total}")

# 4. Update the Toggl entry
print("\n4. Updating Toggl entry...")
if entry_to_update:
    entry_id = entry_to_update['id']
    update_payload = {
        'description': f'{total}x الفاتحة',
    }
    if project_id:
        update_payload['project_id'] = int(project_id)
    
    req = request.Request(
        f'https://api.track.toggl.com/api/v9/workspaces/{workspace_id}/time_entries/{entry_id}',
        data=json.dumps(update_payload).encode(),
        headers={
            'Authorization': f'Basic {creds}',
            'Content-Type': 'application/json'
        },
        method='PUT'
    )
    with request.urlopen(req) as resp:
        result = json.loads(resp.read().decode())
        print(f"   Updated entry {entry_id} to: {result.get('description')}")

print("\n✓ Done!")
