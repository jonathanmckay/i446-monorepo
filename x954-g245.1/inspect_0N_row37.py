import pandas as pd
from openpyxl import load_workbook

def inspect_0N_row37():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0₦']
    
    print("--- 0₦ Sheet Rows 36-38 ---")
    data = []
    # Header row 1 for context
    headers = ["Row"]
    for c in range(1, 20):
        headers.append(str(sheet.cell(row=1, column=c).value))
        
    for r in range(36, 39):
        row_data = [f"R{r}"]
        for c in range(1, 20): 
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)

    df = pd.DataFrame(data, columns=headers)
    print(df.to_markdown(index=False))

if __name__ == "__main__":
    inspect_0N_row37()
