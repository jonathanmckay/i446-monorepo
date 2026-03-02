import pandas as pd
from openpyxl import load_workbook

def inspect_0N():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    if '0₦' not in wb.sheetnames:
        print("'0₦' sheet not found.")
        return

    sheet = wb['0₦']
    print("--- 0₦ Sheet Inspection ---")
    
    # Check headers and first few rows
    data = []
    for r in range(1, 10):
        row_data = []
        for c in range(1, 20): 
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)

    df = pd.DataFrame(data, columns=[f"C{i}" for i in range(1, 20)])
    print(df.to_markdown(index=False))

    # Check for dates in Col 1 or 2
    print("\n--- Dates check ---")
    for r in range(1, 50):
        c1 = sheet.cell(row=r, column=1).value
        c2 = sheet.cell(row=r, column=2).value
        print(f"Row {r}: C1={c1}, C2={c2}")

if __name__ == "__main__":
    inspect_0N()
