#!/usr/bin/env python3
from openpyxl import load_workbook
import datetime

wb = load_workbook('neon_agent_copy.xlsx', data_only=True)

target_date = datetime.datetime(2026, 1, 31)

print('=' * 60)
print('JAN 31, 2026 - ANALYSIS BY CATEGORY')
print('=' * 60)

for sheet_name in ['0分', '0₦', '0s', 'hcbi']:
    sheet = wb[sheet_name]
    
    # Get headers from row 1
    headers = []
    for c in range(1, min(50, sheet.max_column + 1)):
        h = sheet.cell(row=1, column=c).value
        headers.append(str(h) if h else f'Col{c}')
    
    # Search for Jan 31
    for r in range(2, min(100, sheet.max_row + 1)):
        for c in range(1, min(10, sheet.max_column + 1)):
            val = sheet.cell(row=r, column=c).value
            if val:
                is_jan31 = False
                if isinstance(val, datetime.datetime):
                    is_jan31 = val.date() == target_date.date()
                elif '2026-01-31' in str(val):
                    is_jan31 = True
                
                if is_jan31:
                    print(f'\n--- Sheet: {sheet_name} (Row {r}) ---')
                    for col_idx in range(1, min(40, sheet.max_column + 1)):
                        cell_val = sheet.cell(row=r, column=col_idx).value
                        hdr = headers[col_idx - 1] if col_idx - 1 < len(headers) else f'Col{col_idx}'
                        if cell_val is not None:
                            print(f'  {hdr}: {cell_val}')
                    break

print('\n' + '=' * 60)
