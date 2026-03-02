#!/usr/bin/env python3
from openpyxl import load_workbook
import datetime

wb = load_workbook('neon_agent_copy.xlsx', data_only=True)

target_date = datetime.datetime(2026, 1, 21)

print('=' * 70)
print('JAN 21, 2026 - FULL CATEGORY BREAKDOWN')
print('=' * 70)

sheets_to_check = ['0分', '0₦', '0s', 'hcbi', 'm5x2', 'i9', '1₦+', '1₹+1s', '1g']

for sheet_name in sheets_to_check:
    if sheet_name not in wb.sheetnames:
        continue
    sheet = wb[sheet_name]
    
    # Get headers from multiple rows if needed (some sheets have multi-row headers)
    headers_r1 = []
    headers_r2 = []
    for c in range(1, min(60, sheet.max_column + 1)):
        h1 = sheet.cell(row=1, column=c).value
        h2 = sheet.cell(row=2, column=c).value
        headers_r1.append(str(h1)[:15] if h1 else '')
        headers_r2.append(str(h2)[:15] if h2 else '')
    
    # Search for Jan 31
    found = False
    for r in range(2, min(150, sheet.max_row + 1)):
        for c in range(1, min(15, sheet.max_column + 1)):
            val = sheet.cell(row=r, column=c).value
            if val:
                is_jan31 = False
                if isinstance(val, datetime.datetime):
                    is_jan31 = val.date() == target_date.date()
                elif '2026-01-31' in str(val):
                    is_jan31 = True
                
                if is_jan31:
                    found = True
                    print(f'\n{"="*70}')
                    print(f'SHEET: {sheet_name} | ROW: {r}')
                    print('=' * 70)
                    
                    # Print all non-empty values with headers
                    for col_idx in range(1, min(60, sheet.max_column + 1)):
                        cell_val = sheet.cell(row=r, column=col_idx).value
                        h1 = headers_r1[col_idx - 1] if col_idx - 1 < len(headers_r1) else ''
                        h2 = headers_r2[col_idx - 1] if col_idx - 1 < len(headers_r2) else ''
                        header = h1 if h1 else h2 if h2 else f'Col{col_idx}'
                        
                        if cell_val is not None:
                            print(f'  [{col_idx:2}] {header:20} => {cell_val}')
                    break
        if found:
            break

print('\n' + '=' * 70)
print('END OF JAN 31 ANALYSIS')
print('=' * 70)
