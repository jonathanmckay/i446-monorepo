import pandas as pd
from openpyxl import load_workbook
import datetime

def analyze_0fen():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0分']
    
    print("Searching for 2026 dates in '0分'...")
    found_2026 = False
    
    # Search first 1000 rows for 2026
    for r in range(1, 1000):
        val = sheet.cell(row=r, column=2).value # Date seems to be Col 2 based on R2 inspection
        if isinstance(val, datetime.datetime):
            if val.year == 2026:
                print(f"Found 2026 date at Row {r}: {val}")
                found_2026 = True
                break
    
    if not found_2026:
        print("No 2026 dates found in first 1000 rows of '0分'.")

    # Map headers from Row 1
    print("\n--- Column Mapping (Row 1) ---")
    headers = {}
    for c in range(1, 30):
        val = sheet.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c
            print(f"Col {c}: {val}")
            
if __name__ == "__main__":
    analyze_0fen()
