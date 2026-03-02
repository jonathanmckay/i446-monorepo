import json
import datetime
from openpyxl import load_workbook

def generate_map():
    print("Loading workbook (data_only=True)...")
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    if '0s' in wb.sheetnames:
        sheet = wb['0s']
        date_map = {}
        print("Scanning '0s' for dates...")
        for r in range(1, 2000): # Scan sufficient range
            val = sheet.cell(row=r, column=2).value
            
            d_str = None
            if isinstance(val, datetime.datetime):
                d_str = val.strftime("%Y-%m-%d")
            elif isinstance(val, str):
                # Try parse
                try:
                    dt = datetime.datetime.strptime(val, "%Y-%m-%d")
                    d_str = dt.strftime("%Y-%m-%d")
                except:
                    pass
            
            if d_str:
                date_map[d_str] = r
        
        print(f"Found {len(date_map)} dates.")
        with open('neon_server/date_map.json', 'w') as f:
            json.dump(date_map, f, indent=2)
        print("Saved to neon_server/date_map.json")
    else:
        print("'0s' sheet not found.")

if __name__ == "__main__":
    generate_map()
