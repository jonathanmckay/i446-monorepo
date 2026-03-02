import pandas as pd
from openpyxl import load_workbook

def inspect_0fen_detail():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=False) # Check formulas too
    sheet = wb['0分']
    
    print("--- 0分 Sheet Content Inspection ---")
    
    # Check Row 1-20
    data = []
    for r in range(1, 20):
        row_data = []
        for c in range(1, 15): 
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)

    df = pd.DataFrame(data, columns=[f"C{i}" for i in range(1, 15)])
    print(df.to_markdown())

    print("\n--- Rows 500-520 (Mid check) ---")
    data = []
    for r in range(500, 520):
        row_data = []
        for c in range(1, 15): 
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)
    df = pd.DataFrame(data, columns=[f"C{i}" for i in range(1, 15)])
    print(df.to_markdown())

if __name__ == "__main__":
    inspect_0fen_detail()
