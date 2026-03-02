import pandas as pd
from openpyxl import load_workbook
import datetime

def deep_inspect_0fen():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0分']
    
    ranges = [(50, 60), (360, 380)]
    
    for start, end in ranges:
        print(f"\n--- Checking Rows {start}-{end} ---")
        for r in range(start, end):
            val = sheet.cell(row=r, column=2).value
            print(f"Row {r}: {val} (Type: {type(val)})")

if __name__ == "__main__":
    deep_inspect_0fen()
