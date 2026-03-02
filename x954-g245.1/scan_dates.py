import pandas as pd
from openpyxl import load_workbook
import datetime

def scan_all_dates():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0分']
    
    dates = []
    print("Scanning Col 2 for dates...")
    for r in range(1, 5000):
        val = sheet.cell(row=r, column=2).value
        valid_date = False
        if isinstance(val, datetime.datetime):
            dates.append((r, val.date()))
            valid_date = True
        elif val is not None:
             # Try parsing string if needed, but openpyxl usually handles it
             pass
    
    if dates:
        print(f"Found {len(dates)} dates.")
        print(f"First: {dates[0]}")
        print(f"Last: {dates[-1]}")
        # Print last few
        for d in dates[-5:]:
            print(d)
    else:
        print("No dates found.")

if __name__ == "__main__":
    scan_all_dates()
