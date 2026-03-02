from openpyxl import load_workbook
import datetime

def debug_row():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0s']
    
    val = sheet.cell(row=38, column=2).value # Row 38 should be Feb 6
    print(f"Row 38 Col 2 Value: {val}")
    print(f"Type: {type(val)}")
    
    if isinstance(val, datetime.datetime):
        print(f"Date: {val.date()}")
        target = datetime.date(2026, 2, 6)
        print(f"Target: {target}")
        print(f"Match? {val.date() == target}")
    else:
        print("Not a datetime")

if __name__ == "__main__":
    debug_row()
