import pandas as pd
from openpyxl import load_workbook

def analyze_sheet_headers():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    # List actual sheet names to confirm '0s' vs '0分'
    print(f"Actual Sheet Names: {wb.sheetnames}")
    
    sheet_name = '0s' # Default assumption
    if '0分' in wb.sheetnames:
        sheet_name = '0分'
    
    print(f"\nAnalyzing Sheet: {sheet_name}")
    sheet = wb[sheet_name]
    
    # Print first 5 rows to identify header rows
    print("\n--- First 5 Rows ---")
    data = []
    for r in range(1, 6):
        row_data = [f"R{r}"]
        for c in range(1, 20):
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else "")
        data.append(row_data)
        
    df = pd.DataFrame(data, columns=["Row"] + [f"C{i}" for i in range(1, 20)])
    print(df.to_markdown(index=False))

    # Try to identify semantic columns
    # Based on previous look:
    # Row 1: "What is my motivation...", "Title of the day"
    # Row 2: "pri"
    # Row 3: "Goal", "分", "Focus Bonus"
    
    print("\n--- Inferred Schema ---")
    # Mapping C index to potential headers in Row 1, 2, 3
    for c in range(1, 25):
        h1 = sheet.cell(row=1, column=c).value
        h2 = sheet.cell(row=2, column=c).value
        h3 = sheet.cell(row=3, column=c).value
        print(f"Col {c}: [R1] {h1} | [R2] {h2} | [R3] {h3}")

if __name__ == "__main__":
    analyze_sheet_headers()
