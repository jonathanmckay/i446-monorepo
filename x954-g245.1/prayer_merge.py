#!/usr/bin/env python3
"""
Prayer Merge Script

Syncs prayer count from Excel to Toggl:
1. Excel (0₦ sheet, col 39 "ص") is the source of truth for prayer counts
2. Creates a single Toggl entry showing that count: "53x الفاتحة"

The script:
- Reads prayer count from Excel for the target date
- Creates ONE Toggl entry with format "Nx الفاتحة"

For manual adjustments, use --add N to append +N to the Excel cell formula.

Environment Variables Required:
- TOGGL_API_TOKEN: Your Toggl API token
- TOGGL_WORKSPACE_ID: Your Toggl workspace ID
- TOGGL_PROJECT_ID: (Optional) Project ID for the prayer entry

Usage:
    python prayer_merge.py [--date YYYY-MM-DD] [--dry-run] [--add N]
    python prayer_merge.py --setup  # Store credentials in system keychain
"""

import argparse
import datetime
import os
import re
import sys
import base64
import json
from typing import Optional, List, Tuple
from urllib import request, error
from openpyxl import load_workbook

# Keyring service name
KEYRING_SERVICE = 'neon-prayer-merge'


def get_credential(key: str) -> Optional[str]:
    """Get a credential from keychain, falling back to .env file.
    
    Priority:
    1. Environment variable (for CI/GitHub Actions)
    2. System keychain (syncs across devices via iCloud)
    3. Local .env file (fallback)
    """
    # Check environment first (for CI)
    if key in os.environ:
        return os.environ[key]
    
    # Try keychain
    try:
        import keyring
        value = keyring.get_password(KEYRING_SERVICE, key)
        if value:
            return value
    except ImportError:
        pass
    except Exception:
        pass
    
    # Fallback to .env file
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    if k == key and v:
                        return v
    
    return None


def set_credential(key: str, value: str) -> bool:
    """Store a credential in the system keychain."""
    try:
        import keyring
        keyring.set_password(KEYRING_SERVICE, key, value)
        return True
    except ImportError:
        print("Error: keyring not installed. Run: pip install keyring")
        return False
    except Exception as e:
        print(f"Error storing credential: {e}")
        return False


def setup_credentials():
    """Interactive setup to store Toggl credentials in keychain."""
    print("Toggl Credential Setup")
    print("=" * 40)
    print("Credentials will be stored in your system keychain")
    print("(syncs via iCloud Keychain on macOS)")
    print()
    
    # Get current values
    current_token = get_credential('TOGGL_API_TOKEN')
    current_workspace = get_credential('TOGGL_WORKSPACE_ID')
    current_project = get_credential('TOGGL_PROJECT_ID')
    
    # API Token
    if current_token:
        print(f"Current API Token: {current_token[:8]}...{current_token[-4:]}")
        token = input("New API Token (or Enter to keep): ").strip()
        if not token:
            token = current_token
    else:
        token = input("Toggl API Token: ").strip()
    
    if not token:
        print("Error: API Token is required")
        return False
    
    # Workspace ID - try to auto-detect
    if not current_workspace:
        print("\nFetching workspace ID from Toggl...")
        try:
            creds = base64.b64encode(f'{token}:api_token'.encode()).decode()
            req = request.Request(
                'https://api.track.toggl.com/api/v9/me',
                headers={'Authorization': f'Basic {creds}'}
            )
            with request.urlopen(req) as resp:
                data = json.loads(resp.read().decode())
                current_workspace = str(data.get('default_workspace_id', ''))
                print(f"Found workspace: {current_workspace}")
        except Exception as e:
            print(f"Could not auto-detect: {e}")
    
    if current_workspace:
        print(f"Current Workspace ID: {current_workspace}")
        workspace = input("New Workspace ID (or Enter to keep): ").strip()
        if not workspace:
            workspace = current_workspace
    else:
        workspace = input("Toggl Workspace ID: ").strip()
    
    # Project ID (optional)
    if current_project:
        print(f"Current Project ID: {current_project}")
    project = input("Toggl Project ID (optional, Enter to skip): ").strip()
    if not project and current_project:
        project = current_project
    
    # Store credentials
    print("\nStoring credentials in keychain...")
    
    if not set_credential('TOGGL_API_TOKEN', token):
        return False
    print("  ✓ API Token")
    
    if workspace:
        if not set_credential('TOGGL_WORKSPACE_ID', workspace):
            return False
        print("  ✓ Workspace ID")
    
    if project:
        if not set_credential('TOGGL_PROJECT_ID', project):
            return False
        print("  ✓ Project ID")
    
    print("\n✓ Credentials stored successfully!")
    print("  They will sync across your devices via iCloud Keychain.")
    return True


# Constants
PRAYER_COL = 39  # ص column in 0₦ sheet
DATE_COL = 3     # Date column in 0₦ sheet
SHEET_NAME = '0₦'
EXCEL_FILE = 'neon_agent_copy.xlsx'

# Toggl API
TOGGL_API_BASE = 'https://api.track.toggl.com/api/v9'

# Pattern to match prayer entries like "10x الفاتحة" or "الفاتحة 10x" or "10 الفاتحة"
PRAYER_PATTERN = re.compile(r'(\d+)\s*[xX×]?\s*الفاتحة|الفاتحة\s*[xX×]?\s*(\d+)')


def get_toggl_auth_header() -> Optional[str]:
    """Get the Authorization header for Toggl API."""
    api_token = get_credential('TOGGL_API_TOKEN')
    if not api_token:
        return None
    credentials = base64.b64encode(f"{api_token}:api_token".encode()).decode()
    return f'Basic {credentials}'


def get_toggl_prayers(target_date: datetime.date) -> Tuple[int, List[dict]]:
    """Fetch prayer entries from Toggl for the target date.
    
    Searches for time entries with descriptions matching prayer patterns
    (e.g., "10x الفاتحة", "الفاتحة", etc.)
    
    Args:
        target_date: Date to search for
        
    Returns:
        Tuple of (total_count, list_of_entries)
    """
    auth_header = get_toggl_auth_header()
    if not auth_header:
        print("Warning: TOGGL_API_TOKEN not set, skipping Toggl fetch")
        return 0, []
    
    # Toggl API uses start_date and end_date for filtering
    # We need to get entries for the full day
    start_date = target_date.strftime('%Y-%m-%d')
    end_date = (target_date + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    
    url = f"{TOGGL_API_BASE}/me/time_entries?start_date={start_date}&end_date={end_date}"
    
    headers = {
        'Content-Type': 'application/json',
        'Authorization': auth_header
    }
    
    try:
        req = request.Request(url, headers=headers, method='GET')
        with request.urlopen(req) as response:
            if response.status == 200:
                entries = json.loads(response.read().decode())
                return _parse_prayer_entries(entries)
            else:
                print(f"Warning: Toggl API returned status {response.status}")
                return 0, []
                
    except error.HTTPError as e:
        print(f"Warning: Error fetching Toggl entries: {e.code} {e.reason}")
        return 0, []
    except error.URLError as e:
        print(f"Warning: Error connecting to Toggl: {e.reason}")
        return 0, []


def _parse_prayer_entries(entries: List[dict]) -> Tuple[int, List[dict]]:
    """Parse Toggl entries and extract prayer counts.
    
    Args:
        entries: List of Toggl time entry objects
        
    Returns:
        Tuple of (total_count, list_of_matching_entries)
    """
    total_count = 0
    prayer_entries = []
    
    for entry in entries:
        description = entry.get('description', '')
        if not description:
            continue
            
        # Check if this is a prayer entry
        if 'الفاتحة' in description:
            prayer_entries.append(entry)
            
            # Try to extract count from description
            match = PRAYER_PATTERN.search(description)
            if match:
                # Get the count from whichever group matched
                count_str = match.group(1) or match.group(2)
                if count_str:
                    total_count += int(count_str)
            else:
                # If no count specified, assume 1
                total_count += 1
    
    return total_count, prayer_entries


def delete_toggl_entry(entry_id: int, dry_run: bool = False) -> bool:
    """Delete a Toggl time entry.
    
    Args:
        entry_id: The ID of the entry to delete
        dry_run: If True, don't actually delete
        
    Returns:
        bool: True if successful
    """
    if dry_run:
        print(f"  [DRY RUN] Would delete entry {entry_id}")
        return True
        
    auth_header = get_toggl_auth_header()
    if not auth_header:
        return False
    
    workspace_id = get_credential('TOGGL_WORKSPACE_ID')
    if not workspace_id:
        return False
    
    url = f"{TOGGL_API_BASE}/workspaces/{workspace_id}/time_entries/{entry_id}"
    
    headers = {
        'Authorization': auth_header
    }
    
    try:
        req = request.Request(url, headers=headers, method='DELETE')
        with request.urlopen(req) as response:
            return response.status in (200, 204)
    except error.HTTPError as e:
        print(f"Warning: Error deleting entry {entry_id}: {e.code}")
        return False
    except error.URLError as e:
        print(f"Warning: Error connecting to Toggl: {e.reason}")
        return False


def find_date_row(filepath: str, target_date: datetime.date) -> Optional[int]:
    """Find the row number for a given date in the Excel file.
    
    The 0₦ sheet uses formulas for dates (=C5+1, etc.) which openpyxl can't evaluate.
    We use a known anchor date (Row 5 = 2026-01-04) and calculate the offset.
    
    Args:
        filepath: Path to xlsx file
        target_date: datetime.date object
        
    Returns:
        int: Row number, or None if out of range
    """
    # Anchor: Row 5 = 2026-01-04
    ANCHOR_ROW = 5
    ANCHOR_DATE = datetime.date(2026, 1, 4)
    
    # Calculate offset
    days_diff = (target_date - ANCHOR_DATE).days
    target_row = ANCHOR_ROW + days_diff
    
    # Sanity check - don't go before row 5 or too far ahead
    if target_row < 5 or target_row > 500:
        return None
    
    return target_row


def get_prayer_count_from_excel(filepath: str, target_date: datetime.date) -> Tuple[Optional[int], Optional[str]]:
    """Get prayer count (الفاتحة) for a specific date from the Excel file.
    
    Args:
        filepath: Path to xlsx file
        target_date: datetime.date object
        
    Returns:
        Tuple of (evaluated_count, cell_formula_or_value)
        - evaluated_count: The calculated value (int), or None if not found
        - cell_formula: The raw cell content (formula string or value)
    """
    if not os.path.exists(filepath):
        print(f"Error: Excel file not found: {filepath}")
        return None, None
    
    row = find_date_row(filepath, target_date)
    if row is None:
        return None, None
    
    # Get evaluated value (data_only=True)
    wb_data = load_workbook(filepath, data_only=True)
    if SHEET_NAME not in wb_data.sheetnames:
        wb_data.close()
        return None, None
    
    sheet_data = wb_data[SHEET_NAME]
    prayer_val = sheet_data.cell(row=row, column=PRAYER_COL).value
    wb_data.close()
    
    evaluated = 0
    if prayer_val is not None:
        evaluated = int(prayer_val) if isinstance(prayer_val, (int, float)) else 0
    
    # Get formula/raw value (data_only=False)
    wb_formula = load_workbook(filepath, data_only=False)
    sheet_formula = wb_formula[SHEET_NAME]
    cell_content = sheet_formula.cell(row=row, column=PRAYER_COL).value
    wb_formula.close()
    
    return evaluated, cell_content


def append_to_excel_cell(filepath: str, target_date: datetime.date, value_to_add: int, dry_run: bool = False) -> bool:
    """Append a value to the prayer cell formula for debugability.
    
    If the cell contains a formula like "=85", it becomes "=85+15".
    If the cell contains a plain value like 85, it becomes "=85+15".
    If the cell is empty, it becomes "=15".
    
    Args:
        filepath: Path to xlsx file
        target_date: Date to update
        value_to_add: Integer value to append (e.g., 15 becomes "+15")
        dry_run: If True, don't actually modify the file
        
    Returns:
        bool: True if successful
    """
    if value_to_add == 0:
        return True  # Nothing to add
    
    row = find_date_row(filepath, target_date)
    if row is None:
        print(f"Error: Could not find row for {target_date}")
        return False
    
    # Load workbook preserving formulas
    wb = load_workbook(filepath, data_only=False)
    
    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: Sheet '{SHEET_NAME}' not found")
        wb.close()
        return False
    
    sheet = wb[SHEET_NAME]
    cell = sheet.cell(row=row, column=PRAYER_COL)
    current_value = cell.value
    
    # Build new formula
    if current_value is None or current_value == '':
        # Empty cell - start fresh
        new_value = f"={value_to_add}"
    elif isinstance(current_value, str) and current_value.startswith('='):
        # Already a formula - append to it
        new_value = f"{current_value}+{value_to_add}"
    else:
        # Plain value - convert to formula and append
        new_value = f"={current_value}+{value_to_add}"
    
    if dry_run:
        print(f"  [DRY RUN] Would update Excel cell:")
        print(f"    Row {row}, Col {PRAYER_COL} (ص)")
        print(f"    Current: {current_value}")
        print(f"    New:     {new_value}")
        wb.close()
        return True
    
    # Update the cell
    cell.value = new_value
    wb.save(filepath)
    wb.close()
    
    print(f"  Updated Excel: {current_value} → {new_value}")
    return True


def create_toggl_entry(prayer_count: int, target_date: datetime.date, dry_run: bool = False) -> bool:
    """Create a Toggl time entry for the prayer count.
    
    Creates an entry with description "Nx الفاتحة" where N is the prayer count.
    The entry is created with 1 second duration (symbolic entry).
    
    Args:
        prayer_count: Number of prayers
        target_date: Date for the entry
        dry_run: If True, don't actually create the entry
        
    Returns:
        bool: True if successful
    """
    api_token = get_credential('TOGGL_API_TOKEN')
    workspace_id = get_credential('TOGGL_WORKSPACE_ID')
    project_id = get_credential('TOGGL_PROJECT_ID')
    
    if not api_token:
        print("Error: TOGGL_API_TOKEN not found. Run: python prayer_merge.py --setup")
        return False
    
    if not workspace_id:
        print("Error: TOGGL_WORKSPACE_ID not found. Run: python prayer_merge.py --setup")
        return False
    
    description = f"{prayer_count}x الفاتحة"
    
    # Create entry at 5:30 AM on the target date
    start_time = datetime.datetime.combine(
        target_date, 
        datetime.time(5, 30, 0)
    )
    
    # Use ISO format with timezone
    start_iso = start_time.strftime('%Y-%m-%dT%H:%M:%S') + '+00:00'
    
    payload = {
        'description': description,
        'workspace_id': int(workspace_id),
        'start': start_iso,
        'duration': 1,  # 1 second (symbolic)
        'created_with': 'neon_prayer_merge'
    }
    
    if project_id:
        payload['project_id'] = int(project_id)
    
    if dry_run:
        print(f"[DRY RUN] Would create Toggl entry:")
        print(f"  Description: {description}")
        print(f"  Date: {target_date}")
        print(f"  Workspace: {workspace_id}")
        if project_id:
            print(f"  Project: {project_id}")
        return True
    
    # Make API request
    url = f"{TOGGL_API_BASE}/workspaces/{workspace_id}/time_entries"
    
    credentials = base64.b64encode(f"{api_token}:api_token".encode()).decode()
    
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Basic {credentials}'
    }
    
    data = json.dumps(payload).encode('utf-8')
    
    try:
        req = request.Request(url, data=data, headers=headers, method='POST')
        with request.urlopen(req) as response:
            if response.status in (200, 201):
                result = json.loads(response.read().decode())
                print(f"✓ Created Toggl entry: {description}")
                print(f"  Entry ID: {result.get('id')}")
                return True
            else:
                print(f"Error: Unexpected status {response.status}")
                return False
                
    except error.HTTPError as e:
        print(f"Error creating Toggl entry: {e.code} {e.reason}")
        try:
            error_body = e.read().decode()
            print(f"  Response: {error_body}")
        except:
            pass
        return False
    except error.URLError as e:
        print(f"Error connecting to Toggl: {e.reason}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Sync prayer count from Excel to Toggl'
    )
    parser.add_argument(
        '--setup',
        action='store_true',
        help='Store Toggl credentials in system keychain'
    )
    parser.add_argument(
        '--date',
        help='Date to process (YYYY-MM-DD). Default: yesterday'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Print what would be done without making changes'
    )
    parser.add_argument(
        '--excel-file',
        default=EXCEL_FILE,
        help=f'Path to Excel file (default: {EXCEL_FILE})'
    )
    parser.add_argument(
        '--add',
        type=int,
        default=0,
        help='Add N prayers to the Excel cell (appends +N to formula)'
    )
    
    args = parser.parse_args()
    
    # Handle --setup
    if args.setup:
        success = setup_credentials()
        sys.exit(0 if success else 1)
    
    # Determine target date
    if args.date:
        try:
            target_date = datetime.datetime.strptime(args.date, '%Y-%m-%d').date()
        except ValueError:
            print(f"Error: Invalid date format '{args.date}'. Use YYYY-MM-DD")
            sys.exit(1)
    else:
        target_date = datetime.date.today() - datetime.timedelta(days=1)
    
    print(f"Prayer Merge for {target_date}")
    print("=" * 50)
    
    # Step 1: Get current prayer count from Excel
    print("\n📖 Reading Excel (0₦ sheet, col 39)...")
    excel_count, excel_formula = get_prayer_count_from_excel(args.excel_file, target_date)
    
    if excel_count is None:
        print(f"  Warning: Date {target_date} not found in Excel")
        excel_count = 0
        excel_formula = None
    else:
        print(f"  Cell content: {excel_formula}")
        print(f"  Current count: {excel_count}")
    
    # Step 2: If --add specified, append to Excel
    if args.add > 0:
        print(f"\n📝 Adding +{args.add} to Excel cell...")
        if not append_to_excel_cell(args.excel_file, target_date, args.add, dry_run=args.dry_run):
            print("  Error: Failed to update Excel cell")
            sys.exit(1)
        # Add to our count (since openpyxl can't evaluate the formula)
        excel_count += args.add
        print(f"  New total: {excel_count}")
    
    if excel_count == 0:
        print("\n⚠️  No prayers to record.")
        sys.exit(0)
    
    # Step 3: Create Toggl entry with the Excel count
    print("\n✨ Creating Toggl entry...")
    success = create_toggl_entry(excel_count, target_date, dry_run=args.dry_run)
    
    if not success:
        sys.exit(1)
    
    print("\n" + "=" * 50)
    print("Prayer merge complete!")


if __name__ == '__main__':
    main()
