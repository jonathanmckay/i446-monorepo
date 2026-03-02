import pandas as pd
from openpyxl import load_workbook

def inspect_0N_cols30_40():
    wb = load_workbook('neon_agent_copy.xlsx', data_only=True)
    sheet = wb['0₦']
    
    print("--- 0₦ Sheet Rows 30-35, Cols 30-40 ---")
    data = []
    # Header row 1
    headers = ["Row"]
    for c in range(30, 41):
        val = sheet.cell(row=1, column=c).value
        headers.append(f"C{c}:{str(val)}")
        
    for r in range(30, 36):
        row_data = [f"R{r}"]
        for c in range(30, 41): 
            val = sheet.cell(row=r, column=c).value
            row_data.append(str(val) if val is not None else ".")
        data.append(row_data)

    df = pd.DataFrame(data, columns=headers)
    print(df.to_markdown(index=False))

if __name__ == "__main__":
    inspect_0N_cols30_40()
